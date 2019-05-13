import openpyxl

class majorcollege2dict:
    def __init__(self,mappingpath):
        self.mappingpath=mappingpath

    #专业为key,学院为value
    def major_college_mapping(self):
        data = openpyxl.load_workbook(self.mappingpath)
        sheet1 = data['sheet1']
        if sheet1.max_row<2:
            print("学院专业映射文件行数少于2,文件路径为:{}".format(self.mappingpath))
            raise Exception("学院专业映射文件行数少于2,文件路径为:{}".format(self.mappingpath))
        if sheet1.max_column<2:
            print("学院专业映射文件列数少于2,文件路径为:{}".format(self.mappingpath))
            raise Exception("学院专业映射文件列数少于2,文件路径为:{}".format(self.mappingpath))
        major_college_mapper={}
        for i in range(2,sheet1.max_row):
            major=sheet1.cell(row=i,column=2).value
            college=sheet1.cell(row=i,column=1).value
            status=sheet1.cell(row=i,column=3).value
            major_college_mapper[major]={'college':college,'status':status}
        print(major_college_mapper)
        return major_college_mapper
    #学院为key,专业列表为value,
    def college_major_mapping(self):
        data = openpyxl.load_workbook(self.mappingpath)
        sheet1 = data['sheet1']
        if sheet1.max_row<2:
            print("学院专业映射文件行数少于2,文件路径为:{}".format(self.mappingpath))
            raise Exception("学院专业映射文件行数少于2,文件路径为:{}".format(self.mappingpath))
        if sheet1.max_column<2:
            print("学院专业映射文件列数少于2,文件路径为:{}".format(self.mappingpath))
            raise Exception("学院专业映射文件列数少于2,文件路径为:{}".format(self.mappingpath))
        college_major_mapper={}
        for i in range(2,sheet1.max_row):
            major=sheet1.cell(row=i,column=2).value
            college=sheet1.cell(row=i,column=1).value
            status=sheet1.cell(row=i,column=3).value
            obj={'major':major,'status':status}
            if college in college_major_mapper:
                college_major_mapper[college].append(obj)
            else:
                college_major_mapper[college]=[]
                college_major_mapper[college].append(obj)
        print(college_major_mapper)
        return college_major_mapper
    #设置专业的处理状态
    def set_major_status(self,major_name,status=1):
        data = openpyxl.load_workbook(self.mappingpath)
        sheet1 = data['sheet1']
        for i in range(2,sheet1.max_row):
            major=sheet1.cell(row=i,column=2).value
            if major==major_name:
                sheet1.cell(row=i,column=3).value=status
                data.save(self.mappingpath)

        print(major_name+"状态设置为{}".format(status))

if __name__ == "__main__":
    #test_major_college_mapping()
    path=r'e:\newjincin\projects\ros\doc\refresh\datasource\18届数据\院系-专业对照表.xlsx'
    obj=majorcollege2dict(path)
    print(obj.major_college_mapping())
    print(obj.college_major_mapping())
    obj.set_major_status('英语',status=0)